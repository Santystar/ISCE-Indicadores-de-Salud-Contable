import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from config.rutas import obtener_ruta_salida


# ==============================
# FUNCIONES AUXILIARES
# ==============================

def convertir_celda_a_fila_columna(celda):
    """
    Convierte una celda tipo 'A4' en fila y columna numérica.
    """
    letras = ''.join([c for c in celda if c.isalpha()])
    numeros = ''.join([c for c in celda if c.isdigit()])

    columna = 0
    for letra in letras.upper():
        columna = columna * 26 + (ord(letra) - ord('A') + 1)

    fila = int(numeros)

    return fila, columna


def ajustar_ancho_columnas(hoja, columna_inicio, cantidad_columnas):
    """
    Ajusta el ancho de las columnas según el contenido.
    """
    for col in range(columna_inicio, columna_inicio + cantidad_columnas):
        letra_columna = get_column_letter(col)
        ancho_maximo = 12

        for celda in hoja[letra_columna]:
            if celda.value is not None:
                largo = len(str(celda.value))
                if largo > ancho_maximo:
                    ancho_maximo = largo

        hoja.column_dimensions[letra_columna].width = min(ancho_maximo + 2, 40)


# ==============================
# FUNCIÓN PRINCIPAL DE ESCRITURA
# ==============================

def escribir_dataframe_en_excel(
    df,
    nombre_archivo,
    nombre_hoja,
    celda_inicio,
    columna_porcentaje=None,
    formato_porcentaje='0.0%',
    columnas_fecha=None,
    calcular_promedio=False
):
    """
    Escribe un DataFrame en un archivo Excel sin borrar lo anterior.
    """

    ruta_archivo = obtener_ruta_salida(nombre_archivo)

    # ==============================
    # CREAR ARCHIVO SI NO EXISTE
    # ==============================
    if not os.path.exists(ruta_archivo):
        wb = Workbook()
        ws = wb.active
        ws.title = nombre_hoja
        wb.save(ruta_archivo)

    # ==============================
    # CARGAR ARCHIVO EXISTENTE
    # ==============================
    wb = load_workbook(ruta_archivo)

    if nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
    else:
        hoja = wb.create_sheet(title=nombre_hoja)

    fila_inicio, columna_inicio = convertir_celda_a_fila_columna(celda_inicio)

    # ==============================
    # ESCRIBIR ENCABEZADOS
    # ==============================
    for i, columna in enumerate(df.columns):
        hoja.cell(row=fila_inicio, column=columna_inicio + i, value=columna)

    # ==============================
    # ESCRIBIR DATOS
    # ==============================
    for fila_idx, fila in enumerate(df.itertuples(index=False), start=1):
        for col_idx, valor in enumerate(fila, start=0):
            celda = hoja.cell(
                row=fila_inicio + fila_idx,
                column=columna_inicio + col_idx,
                value=valor
            )

            # Aplicar formato porcentaje
            if columna_porcentaje is not None and col_idx == columna_porcentaje:
                celda.number_format = formato_porcentaje

            # Aplicar formato fecha
            if columnas_fecha and col_idx in columnas_fecha:
                celda.number_format = 'DD/MM/YYYY'

    # ==============================
    # CALCULAR PROMEDIO SI APLICA
    # ==============================
    if calcular_promedio and columna_porcentaje is not None:
        fila_promedio = fila_inicio + len(df) + 1

        columna_texto = columna_inicio + columna_porcentaje - 1
        columna_valor = columna_inicio + columna_porcentaje

        hoja.cell(row=fila_promedio, column=columna_texto, value="Promedio")

        rango_inicio = fila_inicio + 1
        rango_fin = fila_inicio + len(df)

        letra_columna = get_column_letter(columna_valor)

        formula = f"=AVERAGE({letra_columna}{rango_inicio}:{letra_columna}{rango_fin})"

        celda_promedio = hoja.cell(row=fila_promedio, column=columna_valor, value=formula)
        celda_promedio.number_format = formato_porcentaje

    # ==============================
    # AJUSTAR ANCHO
    # ==============================
    ajustar_ancho_columnas(hoja, columna_inicio, len(df.columns))

    # ==============================
    # GUARDAR
    # ==============================
    wb.save(ruta_archivo)