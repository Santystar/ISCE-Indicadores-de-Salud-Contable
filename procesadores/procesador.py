import pandas as pd
from openpyxl import load_workbook

from config.configuracion import (
    NOMBRE_BASE_ALCON,
    NOMBRE_BASE_CERTIFICACION,
    NOMBRE_BASE_TEMPORALES,
    NOMBRE_ARCHIVO_SALIDA,
    MES_TRABAJO,
    HOJA_ALCON,
    HOJA_TEMPORAL_REFERENCIA,
    HOJA_SABANA_REFERENCIA,
    COLUMNAS_ALCON,
    COLUMNAS_CERTIFICACION,
    COLUMNAS_TEMPORAL,
    COLUMNAS_SABANA,
    CELDA_INICIO_ALCON,
    CELDA_INICIO_CERTIFICACION,
    CELDA_INICIO_TD_SALDO,
    CELDA_INICIO_TD_SABANA,

    # ==============================
    # NUEVO - CXC
    # ==============================
    NOMBRE_BASE_CXC,
    HOJA_CXC_REFERENCIA,
    HOJA_SABANA_CXC_REFERENCIA,
    COLUMNAS_CXC,
    COLUMNAS_SABANA_CXC,
    CELDA_INICIO_CXC_SALDO,
    CELDA_INICIO_CXC_SABANA
)

from cargadores.cargador_excel import (
    cargar_tabla_excel,
    cargar_tabla_por_coincidencia_hoja
)

from exportadores.exportador_excel import escribir_dataframe_en_excel
from config.rutas import obtener_archivo_por_coincidencia, obtener_ruta_salida


# ==============================
# UTILIDADES
# ==============================

def limpiar_gerencias_invalidas(df, columna="Gerencia"):
    """
    Elimina filas donde la gerencia esté vacía, en blanco o inválida.
    """
    df[columna] = df[columna].astype(str).str.strip()

    df = df[
        (df[columna] != "") &
        (df[columna].str.lower() != "nan") &
        (df[columna].str.lower() != "(en blanco)")
    ]

    return df


def convertir_porcentaje(valor):
    """
    Convierte un valor a número decimal para formato porcentaje en Excel.
    """
    try:
        return float(valor)
    except:
        return 0


def normalizar_si_no(valor):
    """
    Normaliza valores tipo SI / SÍ / NO
    """
    return (
        str(valor)
        .strip()
        .upper()
        .replace("SÍ", "SI")
    )


def construir_base_gerencias(df, columna="gerencia_responsable"):
    """
    Construye lista ordenada de gerencias válidas para alinear tablas.
    """
    base = df[[columna]].copy()
    base[columna] = base[columna].astype(str).str.strip()

    base = base[
        (base[columna] != "") &
        (base[columna].str.lower() != "nan") &
        (base[columna].str.lower() != "(en blanco)")
    ]

    base = base.drop_duplicates().reset_index(drop=True)
    return base


# ==============================
# ALCON
# ==============================

def cargar_alcon_con_encabezado_dinamico():
    """
    Busca automáticamente la fila real de encabezado del archivo ALCON.
    """

    ruta_archivo = obtener_archivo_por_coincidencia(NOMBRE_BASE_ALCON)

    df_crudo = pd.read_excel(ruta_archivo, sheet_name=HOJA_ALCON, header=None)

    fila_encabezado = None

    for i in range(len(df_crudo)):
        fila = df_crudo.iloc[i].astype(str).tolist()

        if "Gerencia" in fila and "Cantidad alertas" in fila:
            fila_encabezado = i
            break

    if fila_encabezado is None:
        raise Exception("No se encontró encabezado en ALCON")

    df = pd.read_excel(ruta_archivo, sheet_name=HOJA_ALCON, header=fila_encabezado)

    # Limpiar nombres de columnas
    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.replace("\n", " ", regex=False)
        .str.replace("\r", " ", regex=False)
    )

    # Buscar columnas reales por coincidencia flexible
    columnas_limpias = {col.lower(): col for col in df.columns}
    columnas_reales = []

    for col in COLUMNAS_ALCON:
        col_limpia = col.lower().strip()

        if col_limpia in columnas_limpias:
            columnas_reales.append(columnas_limpias[col_limpia])
        else:
            raise Exception(f"No se encontró la columna: {col}")

    df = df[columnas_reales].copy()
    df.columns = COLUMNAS_ALCON  # Renombrar para que quede uniforme

    return df


def procesar_alcon():
    """
    Procesa archivo ALCON y lo exporta a la hoja del mes.
    """

    df_alcon = cargar_alcon_con_encabezado_dinamico()

    df_alcon = limpiar_gerencias_invalidas(df_alcon, "Gerencia")

    df_alcon["Calidad Gerencia"] = pd.to_numeric(
        df_alcon["Calidad Gerencia"], errors="coerce"
    ).fillna(0)

    escribir_dataframe_en_excel(
        df=df_alcon,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_ALCON,
        columna_porcentaje=4,
        formato_porcentaje='0.0%'
    )

    return df_alcon


# ==============================
# CERTIFICACIÓN GERENTES
# ==============================

def procesar_certificacion_gerentes():
    """
    Procesa el archivo Histórico Indicador Certificación Gerentes
    y lo exporta al archivo de salida.
    """

    df = cargar_tabla_excel(
        parte_nombre_archivo=NOMBRE_BASE_CERTIFICACION,
        columnas=COLUMNAS_CERTIFICACION
    )

    df["GERENCIA"] = df["GERENCIA"].astype(str).str.strip()

    df = df[
        (df["GERENCIA"] != "") &
        (df["GERENCIA"].str.lower() != "nan") &
        (~df["GERENCIA"].str.startswith("*"))
    ]

    df["FECHA CERTIFICACIÓN"] = pd.to_datetime(
        df["FECHA CERTIFICACIÓN"], errors="coerce", dayfirst=True
    )

    df["FECHA OBJETIVO"] = pd.to_datetime(
        df["FECHA OBJETIVO"], errors="coerce", dayfirst=True
    )

    df["INDICADOR"] = df["INDICADOR"].apply(convertir_porcentaje)

    escribir_dataframe_en_excel(
        df=df,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_CERTIFICACION,
        columna_porcentaje=3,
        calcular_promedio=True,
        formato_porcentaje='0.0%',
        columnas_fecha=[1, 2]
    )

    return df


# ==============================
# TEMPORALES - TD SALDO
# ==============================

def procesar_temporales_td_saldo():
    """
    Procesa la hoja TEMPORAL y construye la tabla TD SALDO.
    """

    df = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo=NOMBRE_BASE_TEMPORALES,
        texto_hoja=HOJA_TEMPORAL_REFERENCIA,
        columnas_esperadas=COLUMNAS_TEMPORAL
    )

    df["gerencia_responsable"] = df["gerencia_responsable"].astype(str).str.strip()

    df = df[
        (df["gerencia_responsable"] != "") &
        (df["gerencia_responsable"].str.lower() != "nan")
    ]

    df["SALDO CONTABLE"] = pd.to_numeric(df["SALDO CONTABLE"], errors="coerce").fillna(0)
    df["PARTIDAS FUERA DE POLITICA_y"] = pd.to_numeric(
        df["PARTIDAS FUERA DE POLITICA_y"], errors="coerce"
    ).fillna(0)

    # Eliminar filas con saldo 0
    df = df[df["SALDO CONTABLE"] != 0]

    # Contar cuentas con saldo
    df["total"] = 1

    # Contar cuántas cuentas tienen fuera de política > 0
    df["fuera"] = (df["PARTIDAS FUERA DE POLITICA_y"] > 0).astype(int)

    resumen = df.groupby("gerencia_responsable", as_index=False).agg({
        "total": "sum",
        "fuera": "sum"
    })

    resumen["%"] = resumen["fuera"] / resumen["total"]

    resumen = resumen.rename(columns={
        "gerencia_responsable": "Area",
        "total": "TOTAL CUENTAS TEMPORALES CON SALDO",
        "fuera": "CUENTAS TEMPORALES FUERA DE POLITICA"
    })

    total_b = resumen["TOTAL CUENTAS TEMPORALES CON SALDO"].sum()
    total_c = resumen["CUENTAS TEMPORALES FUERA DE POLITICA"].sum()
    total_d = total_c / total_b if total_b != 0 else 0

    fila_total = pd.DataFrame([{
        "Area": "Total general",
        "TOTAL CUENTAS TEMPORALES CON SALDO": total_b,
        "CUENTAS TEMPORALES FUERA DE POLITICA": total_c,
        "%": total_d
    }])

    resumen = pd.concat([resumen, fila_total], ignore_index=True)

    escribir_dataframe_en_excel(
        df=resumen,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_TD_SALDO,
        columna_porcentaje=3,
        formato_porcentaje='0.0%'
    )

    return resumen


# ==============================
# TEMPORALES - TD SÁBANA
# ==============================

def procesar_temporales_td_sabana():
    """
    Procesa la hoja Sábana Temporales y construye:
    - TD SÁBANA (conteo)
    - DB / CR por valor
    """

    df = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo=NOMBRE_BASE_TEMPORALES,
        texto_hoja=HOJA_SABANA_REFERENCIA,
        columnas_esperadas=COLUMNAS_SABANA
    )

    df["gerencia_responsable"] = df["gerencia_responsable"].astype(str).str.strip()

    base_gerencias = construir_base_gerencias(df, "gerencia_responsable")

    df["FUERA DE POLITICA"] = df["FUERA DE POLITICA"].apply(normalizar_si_no)
    df["VALOR PARTIDA PESOS"] = pd.to_numeric(df["VALOR PARTIDA PESOS"], errors="coerce").fillna(0)

    # ==============================
    # TABLA E-G (CONTEO DE PARTIDAS)
    # ==============================
    conteo = df.copy()
    conteo["total"] = 1
    conteo["fuera"] = (conteo["FUERA DE POLITICA"] == "SI").astype(int)

    resumen = conteo.groupby("gerencia_responsable", as_index=False).agg({
        "total": "sum",
        "fuera": "sum"
    })

    resumen = base_gerencias.merge(resumen, on="gerencia_responsable", how="left").fillna(0)

    resumen["%"] = resumen["fuera"] / resumen["total"].replace(0, pd.NA)
    resumen["%"] = resumen["%"].fillna(0)

    resumen_export = resumen.rename(columns={
        "total": "TOTAL PARTIDAS",
        "fuera": "PARTIDAS FUERA DE POLITICA"
    })

    resumen_export = resumen_export[[
        "TOTAL PARTIDAS",
        "PARTIDAS FUERA DE POLITICA",
        "%"
    ]]

    total_e = resumen_export["TOTAL PARTIDAS"].sum()
    total_f = resumen_export["PARTIDAS FUERA DE POLITICA"].sum()
    total_g = total_f / total_e if total_e != 0 else 0

    fila_total_eg = pd.DataFrame([{
        "TOTAL PARTIDAS": total_e,
        "PARTIDAS FUERA DE POLITICA": total_f,
        "%": total_g
    }])

    resumen_export = pd.concat([resumen_export, fila_total_eg], ignore_index=True)

    escribir_dataframe_en_excel(
        df=resumen_export,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_TD_SABANA,
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # ==============================
    # TABLA H-J (DÉBITOS)
    # ==============================
    df_db = df[df["VALOR PARTIDA PESOS"] > 0].copy()
    df_db["fuera_db"] = ((df_db["FUERA DE POLITICA"] == "SI") & (df_db["VALOR PARTIDA PESOS"] > 0)).astype(int)

    resumen_db = df_db.groupby("gerencia_responsable", as_index=False).agg({
        "VALOR PARTIDA PESOS": "sum",
        "fuera_db": "sum"
    })

    resumen_db_fuera = df_db[df_db["FUERA DE POLITICA"] == "SI"].groupby(
        "gerencia_responsable", as_index=False
    )["VALOR PARTIDA PESOS"].sum().rename(columns={
        "VALOR PARTIDA PESOS": "VALOR PARTIDAS PESOS DB (FUERA POLITICA)"
    })

    resumen_db = resumen_db.rename(columns={
        "VALOR PARTIDA PESOS": "TOTAL VALOR PARTIDAS PESOS DB"
    })

    resumen_db = base_gerencias.merge(resumen_db, on="gerencia_responsable", how="left")
    resumen_db = resumen_db.merge(resumen_db_fuera, on="gerencia_responsable", how="left")
    resumen_db = resumen_db.fillna(0)

    resumen_db["%"] = resumen_db["VALOR PARTIDAS PESOS DB (FUERA POLITICA)"] / resumen_db["TOTAL VALOR PARTIDAS PESOS DB"].replace(0, pd.NA)
    resumen_db["%"] = resumen_db["%"].fillna(0)

    resumen_db_export = resumen_db[[
        "TOTAL VALOR PARTIDAS PESOS DB",
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)",
        "%"
    ]]

    total_h = resumen_db_export["TOTAL VALOR PARTIDAS PESOS DB"].sum()
    total_i = resumen_db_export["VALOR PARTIDAS PESOS DB (FUERA POLITICA)"].sum()
    total_j = total_i / total_h if total_h != 0 else 0

    fila_total_hj = pd.DataFrame([{
        "TOTAL VALOR PARTIDAS PESOS DB": total_h,
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)": total_i,
        "%": total_j
    }])

    resumen_db_export = pd.concat([resumen_db_export, fila_total_hj], ignore_index=True)

    escribir_dataframe_en_excel(
        df=resumen_db_export,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="H4",
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # ==============================
    # TABLA K-M (CRÉDITOS)
    # ==============================
    df_cr = df[df["VALOR PARTIDA PESOS"] < 0].copy()
    df_cr["VALOR PARTIDA PESOS ABS"] = df_cr["VALOR PARTIDA PESOS"].abs()

    resumen_cr = df_cr.groupby("gerencia_responsable", as_index=False).agg({
        "VALOR PARTIDA PESOS ABS": "sum"
    }).rename(columns={
        "VALOR PARTIDA PESOS ABS": "TOTAL VALOR PARTIDAS PESOS CR"
    })

    resumen_cr_fuera = df_cr[df_cr["FUERA DE POLITICA"] == "SI"].groupby(
        "gerencia_responsable", as_index=False
    )["VALOR PARTIDA PESOS ABS"].sum().rename(columns={
        "VALOR PARTIDA PESOS ABS": "VALOR PARTIDAS PESOS CR (FUERA POLITICA)"
    })

    resumen_cr = base_gerencias.merge(resumen_cr, on="gerencia_responsable", how="left")
    resumen_cr = resumen_cr.merge(resumen_cr_fuera, on="gerencia_responsable", how="left")
    resumen_cr = resumen_cr.fillna(0)

    resumen_cr["%"] = resumen_cr["VALOR PARTIDAS PESOS CR (FUERA POLITICA)"] / resumen_cr["TOTAL VALOR PARTIDAS PESOS CR"].replace(0, pd.NA)
    resumen_cr["%"] = resumen_cr["%"].fillna(0)

    resumen_cr_export = resumen_cr[[
        "TOTAL VALOR PARTIDAS PESOS CR",
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)",
        "%"
    ]]

    total_k = resumen_cr_export["TOTAL VALOR PARTIDAS PESOS CR"].sum()
    total_l = resumen_cr_export["VALOR PARTIDAS PESOS CR (FUERA POLITICA)"].sum()
    total_m = total_l / total_k if total_k != 0 else 0

    fila_total_km = pd.DataFrame([{
        "TOTAL VALOR PARTIDAS PESOS CR": total_k,
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)": total_l,
        "%": total_m
    }])

    resumen_cr_export = pd.concat([resumen_cr_export, fila_total_km], ignore_index=True)

    escribir_dataframe_en_excel(
        df=resumen_cr_export,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="K4",
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # ==============================
    # FORMATO NUMÉRICO Y LIMPIEZA
    # ==============================
    ruta_salida = obtener_ruta_salida(NOMBRE_ARCHIVO_SALIDA)
    wb = load_workbook(ruta_salida)
    ws = wb[MES_TRABAJO]

    ultima_fila = 4 + len(resumen_db_export)

    ultima_fila_cxc = 36 + len(resumen_cr) 

    for fila_excel in range(5, ultima_fila + 1):
        for col in ["H", "I", "K", "L"]:
            ws[f"{col}{fila_excel}"].number_format = '#,##0'

        # TEMPORALES
    for fila_excel in range(5, ultima_fila + 1):
        ws[f"J{fila_excel}"].number_format = '0.0%'
        ws[f"M{fila_excel}"].number_format = '0.0%'

    # CXC (CORRECCIÓN CLAVE)
    for fila_excel in range(37, ultima_fila_cxc + 1):
        ws[f"J{fila_excel}"].number_format = '0.0%'
        ws[f"M{fila_excel}"].number_format = '0.0%'

    for fila_excel in range(4, 1000):
        ws[f"N{fila_excel}"].value = None
        ws[f"N{fila_excel}"].number_format = 'General'

    wb.save(ruta_salida)

    return resumen_export


# ==============================
# CXC
# ==============================

def procesar_cxc():
    """
    Procesa el archivo Informe CxC:
    - Hoja CXC -> A36:D
    - Hoja Sábana CXC -> E36:G
    - Valores DB / CR -> H36:M
    """

    # ==============================
    # HOJA CXC
    # ==============================
    df_cxc = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo=NOMBRE_BASE_CXC,
        texto_hoja=HOJA_CXC_REFERENCIA,
        columnas_esperadas=COLUMNAS_CXC
    )

    df_cxc["gerencia_responsable"] = df_cxc["gerencia_responsable"].astype(str).str.strip()
    df_cxc["SALDO CONTABLE"] = pd.to_numeric(
        df_cxc["SALDO CONTABLE"], errors="coerce"
    ).fillna(0)

    df_cxc["PARTIDAS FUERA DE POLITICA_y"] = pd.to_numeric(
        df_cxc["PARTIDAS FUERA DE POLITICA_y"], errors="coerce"
    ).fillna(0)

    df_cxc = df_cxc[
        (df_cxc["gerencia_responsable"] != "") &
        (df_cxc["gerencia_responsable"].str.lower() != "nan")
    ]

    df_cxc = df_cxc[df_cxc["SALDO CONTABLE"] != 0]

    base_gerencias = construir_base_gerencias(
        df_cxc,
        "gerencia_responsable"
    )

    df_cxc["total"] = 1
    df_cxc["fuera"] = (
        df_cxc["PARTIDAS FUERA DE POLITICA_y"] > 0
    ).astype(int)

    resumen_cxc = df_cxc.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "total": "sum",
        "fuera": "sum"
    })

    resumen_cxc = base_gerencias.merge(
        resumen_cxc,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen_cxc["%"] = (
        resumen_cxc["fuera"] /
        resumen_cxc["total"].replace(0, pd.NA)
    ).fillna(0)

    resumen_cxc_export = resumen_cxc.rename(columns={
        "gerencia_responsable": "Area",
        "total": "TOTAL CUENTAS POR COBRAR CON SALDO",
        "fuera": "CUENTAS POR COBRAR FUERA DE POLITICA"
    })

    resumen_cxc_export = resumen_cxc_export[
        [
            "Area",
            "TOTAL CUENTAS POR COBRAR CON SALDO",
            "CUENTAS POR COBRAR FUERA DE POLITICA",
            "%"
        ]
    ]

    total_b = resumen_cxc_export[
        "TOTAL CUENTAS POR COBRAR CON SALDO"
    ].sum()

    total_c = resumen_cxc_export[
        "CUENTAS POR COBRAR FUERA DE POLITICA"
    ].sum()

    total_d = total_c / total_b if total_b != 0 else 0

    fila_total = pd.DataFrame([{
        "Area": "Total general",
        "TOTAL CUENTAS POR COBRAR CON SALDO": total_b,
        "CUENTAS POR COBRAR FUERA DE POLITICA": total_c,
        "%": total_d
    }])

    resumen_cxc_export = pd.concat(
        [resumen_cxc_export, fila_total],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
        df=resumen_cxc_export,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_CXC_SALDO,
        columna_porcentaje=3,
        formato_porcentaje='0.0%'
    )

    # ==============================
    # HOJA SÁBANA CXC
    # ==============================
    df_sabana = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo=NOMBRE_BASE_CXC,
        texto_hoja=HOJA_SABANA_CXC_REFERENCIA,
        columnas_esperadas=COLUMNAS_SABANA_CXC
    )

    df_sabana["gerencia_responsable"] = df_sabana[
        "gerencia_responsable"
    ].astype(str).str.strip()

    df_sabana["FUERA DE CICLO"] = df_sabana[
        "FUERA DE CICLO"
    ].apply(normalizar_si_no)

    df_sabana["VALOR PARTIDA PESOS"] = pd.to_numeric(
        df_sabana["VALOR PARTIDA PESOS"],
        errors="coerce"
    ).fillna(0)

    df_sabana = df_sabana[
        (df_sabana["gerencia_responsable"] != "") &
        (df_sabana["gerencia_responsable"].str.lower() != "nan")
    ]

    # ==============================
    # E:G CONTEO PARTIDAS
    # ==============================
    df_sabana["total"] = 1
    df_sabana["fuera"] = (
        df_sabana["FUERA DE CICLO"] == "SI"
    ).astype(int)

    resumen_sabana = df_sabana.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "total": "sum",
        "fuera": "sum"
    })

    resumen_sabana = base_gerencias.merge(
        resumen_sabana,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen_sabana["%"] = (
        resumen_sabana["fuera"] /
        resumen_sabana["total"].replace(0, pd.NA)
    ).fillna(0)

    resumen_sabana_export = resumen_sabana.rename(columns={
        "total": "TOTAL PARTIDAS",
        "fuera": "PARTIDAS FUERA DE POLITICA"
    })

    resumen_sabana_export = resumen_sabana_export[
        [
            "TOTAL PARTIDAS",
            "PARTIDAS FUERA DE POLITICA",
            "%"
        ]
    ]

    total_e = resumen_sabana_export["TOTAL PARTIDAS"].sum()
    total_f = resumen_sabana_export["PARTIDAS FUERA DE POLITICA"].sum()
    total_g = total_f / total_e if total_e != 0 else 0

    fila_total_eg = pd.DataFrame([{
        "TOTAL PARTIDAS": total_e,
        "PARTIDAS FUERA DE POLITICA": total_f,
        "%": total_g
    }])

    resumen_sabana_export = pd.concat(
        [resumen_sabana_export, fila_total_eg],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
        df=resumen_sabana_export,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio=CELDA_INICIO_CXC_SABANA,
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # ==============================
    # H:M VALORES DB / CR
    # ==============================
    df_sabana["DB"] = df_sabana[
        "VALOR PARTIDA PESOS"
    ].apply(lambda x: x if x > 0 else 0)

    df_sabana["CR"] = df_sabana[
        "VALOR PARTIDA PESOS"
    ].apply(lambda x: abs(x) if x < 0 else 0)

    df_sabana["DB_fuera"] = df_sabana.apply(
        lambda x: x["DB"] if x["FUERA DE CICLO"] == "SI" else 0,
        axis=1
    )

    df_sabana["CR_fuera"] = df_sabana.apply(
        lambda x: x["CR"] if x["FUERA DE CICLO"] == "SI" else 0,
        axis=1
    )

    resumen_valores = df_sabana.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "DB": "sum",
        "DB_fuera": "sum",
        "CR": "sum",
        "CR_fuera": "sum"
    })

    resumen_valores = base_gerencias.merge(
        resumen_valores,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen_valores["%_db"] = (
        resumen_valores["DB_fuera"] /
        resumen_valores["DB"].replace(0, pd.NA)
    ).fillna(0)

    resumen_valores["%_cr"] = (
        resumen_valores["CR_fuera"] /
        resumen_valores["CR"].replace(0, pd.NA)
    ).fillna(0)

    # ------------------------------
    # DB
    # ------------------------------
    resumen_db = pd.DataFrame({
        "TOTAL VALOR PARTIDAS PESOS DB": resumen_valores["DB"],
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)": resumen_valores["DB_fuera"],
        "%": resumen_valores["%_db"]
    })

    total_h = resumen_db["TOTAL VALOR PARTIDAS PESOS DB"].sum()
    total_i = resumen_db["VALOR PARTIDAS PESOS DB (FUERA POLITICA)"].sum()
    total_j = total_i / total_h if total_h != 0 else 0

    fila_total_db = pd.DataFrame([{
        "TOTAL VALOR PARTIDAS PESOS DB": total_h,
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)": total_i,
        "%": total_j
    }])

    resumen_db = pd.concat(
        [resumen_db, fila_total_db],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
        df=resumen_db,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="H36",
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # ------------------------------
    # CR
    # ------------------------------
    resumen_cr = pd.DataFrame({
        "TOTAL VALOR PARTIDAS PESOS CR": resumen_valores["CR"],
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)": resumen_valores["CR_fuera"],
        "%": resumen_valores["%_cr"]
    })

    total_k = resumen_cr["TOTAL VALOR PARTIDAS PESOS CR"].sum()
    total_l = resumen_cr["VALOR PARTIDAS PESOS CR (FUERA POLITICA)"].sum()
    total_m = total_l / total_k if total_k != 0 else 0

    fila_total_cr = pd.DataFrame([{
        "TOTAL VALOR PARTIDAS PESOS CR": total_k,
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)": total_l,
        "%": total_m
    }])

    resumen_cr = pd.concat(
        [resumen_cr, fila_total_cr],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
        df=resumen_cr,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="K36",
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    return resumen_cxc_export, resumen_sabana_export

# ==============================
# CXP
# ==============================

def procesar_cxp():
    """
    Procesa archivo Informe CxP
    Hoja CXP -> tabla A60:D
    """

    df = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo="Informe CxP",
        texto_hoja="CXP",
        columnas_esperadas=[
            "gerencia_responsable",
            "SALDO CONTABLE",
            "PARTIDAS FUERA DE POLITICA_y"
        ]
    )

    # --------------------------
    # Limpieza base
    # --------------------------
    df["gerencia_responsable"] = (
        df["gerencia_responsable"]
        .astype(str)
        .str.strip()
    )

    df = df[
        (df["gerencia_responsable"] != "") &
        (df["gerencia_responsable"].str.lower() != "nan")
    ]

    df["SALDO CONTABLE"] = pd.to_numeric(
        df["SALDO CONTABLE"],
        errors="coerce"
    ).fillna(0)

    df["PARTIDAS FUERA DE POLITICA_y"] = pd.to_numeric(
        df["PARTIDAS FUERA DE POLITICA_y"],
        errors="coerce"
    ).fillna(0)

    # Solo cuentas con saldo
    df = df[df["SALDO CONTABLE"] != 0]

    # --------------------------
    # Base de gerencias
    # --------------------------
    base_gerencias = construir_base_gerencias(
        df,
        "gerencia_responsable"
    )

    # --------------------------
    # Conteos
    # --------------------------
    df["total"] = 1

    df["fuera"] = (
        df["PARTIDAS FUERA DE POLITICA_y"] > 0
    ).astype(int)

    resumen = df.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "total": "sum",
        "fuera": "sum"
    })

    resumen = base_gerencias.merge(
        resumen,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen["%"] = (
        resumen["fuera"] /
        resumen["total"].replace(0, pd.NA)
    ).fillna(0)

    # --------------------------
    # Exportable
    # --------------------------
    exportar = resumen.rename(columns={
        "gerencia_responsable": "Area",
        "total": "TOTAL CUENTAS POR PAGAR CON SALDO",
        "fuera": "CUENTAS POR PAGAR FUERA DE POLITICA"
    })

    exportar = exportar[[
        "Area",
        "TOTAL CUENTAS POR PAGAR CON SALDO",
        "CUENTAS POR PAGAR FUERA DE POLITICA",
        "%"
    ]]

    # --------------------------
    # Total general
    # --------------------------
    total_b = exportar[
        "TOTAL CUENTAS POR PAGAR CON SALDO"
    ].sum()

    total_c = exportar[
        "CUENTAS POR PAGAR FUERA DE POLITICA"
    ].sum()

    total_d = total_c / total_b if total_b != 0 else 0

    fila_total = pd.DataFrame([{
        "Area": "Total general",
        "TOTAL CUENTAS POR PAGAR CON SALDO": total_b,
        "CUENTAS POR PAGAR FUERA DE POLITICA": total_c,
        "%": total_d
    }])

    exportar = pd.concat(
        [exportar, fila_total],
        ignore_index=True
    )

    # --------------------------
    # Exportar Excel
    # --------------------------
    escribir_dataframe_en_excel(
        df=exportar,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="A60",
        columna_porcentaje=3,
        formato_porcentaje='0.0%'
    )

        # ==================================================
    # SÁBANA CXP  (DB / CR)
    # ==================================================

    df_sabana = cargar_tabla_por_coincidencia_hoja(
        parte_nombre_archivo="Informe CxP",
        texto_hoja="Sábana CXP",
        columnas_esperadas=[
            "gerencia_responsable",
            "FUERA DE CICLO",
            "VALOR PARTIDA PESOS"
        ]
    )

    df_sabana["gerencia_responsable"] = (
        df_sabana["gerencia_responsable"]
        .astype(str)
        .str.strip()
    )

    df_sabana = df_sabana[
        (df_sabana["gerencia_responsable"] != "") &
        (df_sabana["gerencia_responsable"].str.lower() != "nan")
    ]

    df_sabana["VALOR PARTIDA PESOS"] = pd.to_numeric(
        df_sabana["VALOR PARTIDA PESOS"],
        errors="coerce"
    ).fillna(0)

    df_sabana["FUERA DE CICLO"] = (
        df_sabana["FUERA DE CICLO"]
        .astype(str)
        .str.upper()
        .str.strip()
        .replace("SÍ", "SI")
    )

        # =====================================
    # TABLA E:G  TOTAL PARTIDAS
    # =====================================

    df_sabana["total_partidas"] = 1

    df_sabana["fuera_partidas"] = (
        df_sabana["FUERA DE CICLO"] == "SI"
    ).astype(int)

    resumen_partidas = df_sabana.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "total_partidas": "sum",
        "fuera_partidas": "sum"
    })

    resumen_partidas = base_gerencias.merge(
        resumen_partidas,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen_partidas["%"] = (
        resumen_partidas["fuera_partidas"] /
        resumen_partidas["total_partidas"].replace(0, pd.NA)
    ).fillna(0)

    exportar_partidas = pd.DataFrame({
        "TOTAL PARTIDAS":
            resumen_partidas["total_partidas"],
        "PARTIDAS FUERA DE POLITICA":
            resumen_partidas["fuera_partidas"],
        "%":
            resumen_partidas["%"]
    })

    # Total general
    t1 = exportar_partidas.iloc[:,0].sum()
    t2 = exportar_partidas.iloc[:,1].sum()
    t3 = t2 / t1 if t1 != 0 else 0

    fila_total = pd.DataFrame([{
        "TOTAL PARTIDAS": t1,
        "PARTIDAS FUERA DE POLITICA": t2,
        "%": t3
    }])

    exportar_partidas = pd.concat(
        [exportar_partidas, fila_total],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
        df=exportar_partidas,
        nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
        nombre_hoja=MES_TRABAJO,
        celda_inicio="E60",
        columna_porcentaje=2,
        formato_porcentaje='0.0%'
    )

    # -------------------------
    # Separar DB / CR
    # -------------------------
    df_sabana["DB"] = df_sabana["VALOR PARTIDA PESOS"].apply(
        lambda x: x if x > 0 else 0
    )

    df_sabana["CR"] = df_sabana["VALOR PARTIDA PESOS"].apply(
        lambda x: abs(x) if x < 0 else 0
    )

    df_sabana["DB_fuera"] = df_sabana.apply(
        lambda x: x["DB"] if x["FUERA DE CICLO"] == "SI" else 0,
        axis=1
    )

    df_sabana["CR_fuera"] = df_sabana.apply(
        lambda x: x["CR"] if x["FUERA DE CICLO"] == "SI" else 0,
        axis=1
    )

    resumen_valores = df_sabana.groupby(
        "gerencia_responsable",
        as_index=False
    ).agg({
        "DB": "sum",
        "DB_fuera": "sum",
        "CR": "sum",
        "CR_fuera": "sum"
    })

    resumen_valores = base_gerencias.merge(
        resumen_valores,
        on="gerencia_responsable",
        how="left"
    ).fillna(0)

    resumen_valores["%_db"] = (
        resumen_valores["DB_fuera"] /
        resumen_valores["DB"].replace(0, pd.NA)
    ).fillna(0)

    resumen_valores["%_cr"] = (
        resumen_valores["CR_fuera"] /
        resumen_valores["CR"].replace(0, pd.NA)
    ).fillna(0)

    exportar_sabana = pd.DataFrame({
        "TOTAL VALOR PARTIDAS PESOS DB":
            resumen_valores["DB"],
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)":
            resumen_valores["DB_fuera"],
        "%":
            resumen_valores["%_db"],

        "TOTAL VALOR PARTIDAS PESOS CR":
            resumen_valores["CR"],
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)":
            resumen_valores["CR_fuera"],
        "% ":
            resumen_valores["%_cr"]
    })

    # -------------------------
    # Total general
    # -------------------------
    total_db = exportar_sabana.iloc[:,0].sum()
    total_db_fuera = exportar_sabana.iloc[:,1].sum()
    total_db_pct = total_db_fuera / total_db if total_db != 0 else 0

    total_cr = exportar_sabana.iloc[:,3].sum()
    total_cr_fuera = exportar_sabana.iloc[:,4].sum()
    total_cr_pct = total_cr_fuera / total_cr if total_cr != 0 else 0

    fila_total = pd.DataFrame([{
        "TOTAL VALOR PARTIDAS PESOS DB": total_db,
        "VALOR PARTIDAS PESOS DB (FUERA POLITICA)": total_db_fuera,
        "%": total_db_pct,

        "TOTAL VALOR PARTIDAS PESOS CR": total_cr,
        "VALOR PARTIDAS PESOS CR (FUERA POLITICA)": total_cr_fuera,
        "% ": total_cr_pct
    }])

    exportar_sabana = pd.concat(
        [exportar_sabana, fila_total],
        ignore_index=True
    )

    escribir_dataframe_en_excel(
    df=exportar_sabana,
    nombre_archivo=NOMBRE_ARCHIVO_SALIDA,
    nombre_hoja=MES_TRABAJO,
    celda_inicio="H60",
    columnas_porcentaje=[2, 5],
    formato_porcentaje='0.0%'
    )

    return exportar, exportar_sabana